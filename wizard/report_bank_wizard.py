# -*- coding: utf-8 -*-
from odoo import models, fields, _
from odoo.exceptions import ValidationError
import datetime
import logging
import pandas as pd
from openpyxl import load_workbook
logger = logging.getLogger(__name__)


class ReportBank(models.TransientModel):
    _name = 'report.bank'

    file_selection = fields.Selection([('excel', 'EXCEL'),('pdf', 'PDF')], 'Type File', default='excel')
    values_balance_bank = {
        "Bancos": [],
        "Cuentas": [],
        "Estados de Cuentas":[],
        "Ingresos": [],
        "Egresos": [],
        "Saldo Final": [],
    }

    def get_all_balance_banks(self):
        self.env.cr.execute("""
            SELECT rpb.acc_number, rb.name, 
            ab.balance_start, 
            SUM(CASE WHEN absl.amount > 0 THEN absl.amount END) AS income,
            SUM(CASE WHEN absl.amount < 0 THEN absl.amount END) AS egress,
            ab.balance_end_real
            FROM account_bank_statement AS ab
            JOIN account_journal as aj ON aj.id = ab.journal_id
            JOIN res_partner_bank as rpb ON rpb.id = aj.bank_account_id
            JOIN res_bank as rb ON rb.id = rpb.bank_id
            JOIN account_bank_statement_line absl ON absl.statement_id = ab.id
            where aj.type = %s
            and ab.state = %s
            GROUP BY rb.name, rpb.acc_number, ab.balance_start, ab.balance_end_real
            """,('bank', 'confirm'))
        #logger.info(self.env.cr.fetchall())
        return self.env.cr.fetchall()

    def get_total_income_expenses(self):
        self.env.cr.execute("""
            SELECT SUM(CASE WHEN absl.amount > 0 THEN absl.amount END) AS total_income,
            SUM(CASE WHEN absl.amount < 0 THEN absl.amount END) AS total_egress
            FROM account_bank_statement AS ab
            JOIN account_journal as aj ON aj.id = ab.journal_id
            JOIN res_partner_bank as rpb ON rpb.id = aj.bank_account_id
            JOIN account_bank_statement_line absl ON absl.statement_id = ab.id
            where aj.type = %s
            and ab.state = %s
            """,('bank','confirm'))

        return self.env.cr.fetchall()
    
    def get_total_balance_start_end_balance_end(self):
        self.env.cr.execute("""
            SELECT SUM(ab.balance_start), 
            SUM(ab.balance_end_real)
            FROM account_bank_statement AS ab
            INNER JOIN account_journal as aj ON aj.id = ab.journal_id
            where aj.type = %s
            and ab.state = %s
            """,('bank','confirm'))
        return self.env.cr.fetchall()

    def get_date_current(self):
        date = str(datetime.datetime.strptime(str(datetime.datetime.today().date()), "%Y-%m-%d").strftime('%d/%m/%Y'))
        date_only = str(datetime.datetime.strptime(str(datetime.datetime.today().date()), "%Y-%m-%d").strftime('%d/%m/%Y'))
        return date, date_only

    def get_hour_current(self):
        format_hour = str(datetime.datetime.today().strftime("%I:%M %p")).lower()
        return format_hour

    def get_report(self):
        all_balance_banks = self.get_all_balance_banks()
        try:
            if self.file_selection == "excel":
                for balance_bank in all_balance_banks:
                    self.add_data_in_dict(balance_bank)
                    logger.info("PRUEBA")
                return self.create_excel_model_balaance_bank(self.values_balance_bank)
            elif self.file_selection == "pdf":
                return self.get_report_by_pdf()
        except:
            raise ValidationError(_("Surely the data is not complete or there are no records related to bank accounts, please check or contact the system administrator, sorry for the inconvenience caused."))
    
      
    def add_data_in_dict(self, balance_bank):
        self.values_balance_bank["Bancos"].append(balance_bank[0])
        self.values_balance_bank["Cuentas"].append(balance_bank[1])
        self.values_balance_bank["Estados de Cuentas"].append(balance_bank[2])
        self.values_balance_bank["Ingresos"].append(balance_bank[3])
        self.values_balance_bank["Egresos"].append(balance_bank[4])
        self.values_balance_bank["Saldo Final"].append(balance_bank[5])
    
    def create_excel_model_balaance_bank(self, values):
        df = pd.DataFrame(values)
        df = df[["Bancos","Cuentas","Estados de Cuentas","Ingresos", "Egresos", "Saldo Final"]] 
        df.to_excel('balance_bank.xlsx', 'Saldos de Bancos', index=False)
        wb = load_workbook(filename = 'balance_bank.xlsx')
        sheet_ranges = wb['Saldos de Bancos']
        cad = ''
        logger.info('nro de person')
        logger.info(len(values["Bancos"]))
        for row in range(1,len(values["Bancos"])+2):
            for col in range(1,7):
                if str(sheet_ranges.cell(row=row, column=col).value) == "None":
                    cad+=str(' ')+','
                else:
                    cad+=str(sheet_ranges.cell(row=row, column=col).value)+','
            cad+='\n'
        total_egress_income = self.get_total_income_expenses()
        total_balance_start_end = self.get_total_balance_start_end_balance_end()
        cad+=' '+','+'Total'+','+str(total_balance_start_end[0][0])+','+str(total_egress_income[0][0])+','+str(total_egress_income[0][1])+','+str(total_balance_start_end[0][1])+','
        #sheet_ranges.cell(row=1, column=3)
        file_data = cad
        logger.info(file_data)
        values = {
            'name': "reporte_balance_de_cuentas_excel",
            #'datas_fname': 'print_file_name.txt',
            'type': 'binary',
            'res_model': 'ir.ui.view',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'public': True,
            'db_datas': file_data,
            }
        attachment_id = self.env['ir.attachment'].sudo().create(values)
        base_url = self.env['ir.config_parameter'].get_param('web.base.url')
        logger.info(attachment_id)
        download_url = '/web/content/' + str( attachment_id.id, ) + '?download=true'

        return {
                'type' : 'ir.actions.act_url',
                "url": str(base_url) + str(download_url),
                'target': 'new',
            }
    

    def get_report_by_pdf(self):
        hour = str(self.get_hour_current())
        total_bank_balance = self.get_all_balance_banks()
        logger.info(self.get_hour_current())
        total_income_expenses = self.get_total_income_expenses()
        date, date_only = self.get_date_current()
        total_balance_start_end_balance_end = self.get_total_balance_start_end_balance_end()
        logger.info("CAMBIOS")
        logger.info(total_balance_start_end_balance_end)

        data = {
                'form': self.read()[0],
                'bank_balance': total_bank_balance,
                'total_income_expenses': total_income_expenses,
                'total_balance_start_end': total_balance_start_end_balance_end,
                'hours': hour,
                'date':date,
                'date_only':date_only
        }
        
        logger.info(self.env.ref('report_bank.action_report_bank').report_action(self, data=data))
        return self.env.ref('report_bank.action_report_bank').report_action(self, data=data)
        